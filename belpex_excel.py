# -*- coding: utf-8 -*-
"""
belpex_excel.py — bouwt de Excel-werkmap voor de eenheidsprijsberekening.

Twee bladen:

  Kwartierdata  alle kwartierprijzen van de gekozen periode (in EUR/kWh),
                met twee lege gele kolommen: E 'Vergoeding (EUR/kWh)' (bv.
                de marge van je leverancier) en F 'Verbruik (kWh)'. Kolom G
                rekent de kost per kwartier: (prijs + vergoeding) x verbruik.
                Verwacht dat het doorgegeven 'price' al in EUR/kWh staat,
                niet EUR/MWh.

  Overzicht     per maand de verbruiksgewogen eenheidsprijs (incl. vergoeding),
                gesplitst in piek en dal. Alles zijn SUMIFS-formules die naar
                Kwartierdata verwijzen, dus het blad rekent zichzelf zodra er
                verbruik in kolom F staat.

Piek = weekdagen 08:00 t.e.m. 19:45 · Dal = weekdagen 20:00 t.e.m. 07:45 en
alle weekends. Die definitie ligt vast in de werkmap, los van de piekuren
die in de dashboard-sidebar gekozen zijn.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PIEK, DAL = "Piek", "Dal"

BLAUW = "1F4E79"
GEEL = "FFF2CC"
GRIJS = "F2F2F2"

_kop = Font(bold=True, color="FFFFFF", size=11)
_kop_vul = PatternFill("solid", fgColor=BLAUW)
_geel_vul = PatternFill("solid", fgColor=GEEL)
_grijs_vul = PatternFill("solid", fgColor=GRIJS)
_dun = Side(style="thin", color="BFBFBF")
_rand = Border(left=_dun, right=_dun, top=_dun, bottom=_dun)

EUR = '#,##0.00\\ "EUR"'
EUR4 = '#,##0.0000\\ "EUR"'
KWH = '#,##0.000'
DTM = "dd/mm/yyyy hh:mm"


def _is_piek(dt: pd.Series) -> pd.Series:
    """Weekdag tussen 08:00 en 20:00 (20:00 zelf valt al in het dal)."""
    return (dt.dt.dayofweek < 5) & (dt.dt.hour >= 8) & (dt.dt.hour < 20)


def _kopregel(ws, koppen, breedtes, rij=1, hoogte=30):
    for i, (tekst, breedte) in enumerate(zip(koppen, breedtes), start=1):
        cel = ws.cell(row=rij, column=i, value=tekst)
        cel.font = _kop
        cel.fill = _kop_vul
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = _rand
        ws.column_dimensions[get_column_letter(i)].width = breedte
    ws.row_dimensions[rij].height = hoogte


def _blad_kwartierdata(wb: Workbook, d: pd.DataFrame) -> int:
    """Schrijft de kwartierdata en geeft het aantal datarijen terug."""
    ws = wb.create_sheet("Kwartierdata")
    _kopregel(
        ws,
        ["Datum en tijd", "Maand", "Periode", "Prijs (EUR/kWh)",
         "Vergoeding (EUR/kWh)", "Verbruik (kWh)", "Kost (EUR)"],
        [19, 10, 10, 16, 18, 16, 15],
    )

    dt = pd.to_datetime(d["datetime"])
    maand = dt.dt.strftime("%Y-%m")
    periode = _is_piek(dt).map({True: PIEK, False: DAL})
    prijs = pd.to_numeric(d["price"], errors="coerce")
    vbr = (pd.to_numeric(d["verbruik"], errors="coerce")
           if "verbruik" in d.columns else pd.Series([None] * len(d), index=d.index))
    verg = (pd.to_numeric(d["vergoeding"], errors="coerce")
            if "vergoeding" in d.columns else pd.Series([None] * len(d), index=d.index))

    for i, (stempel, mnd, per, prs, vb, vg) in enumerate(
            zip(dt, maand, periode, prijs, vbr, verg), start=2):
        ws.cell(row=i, column=1, value=stempel.to_pydatetime()).number_format = DTM
        ws.cell(row=i, column=2, value=mnd)
        ws.cell(row=i, column=3, value=per)
        ws.cell(row=i, column=4,
                value=None if pd.isna(prs) else float(prs)).number_format = EUR4
        vergoeding = ws.cell(row=i, column=5,
                             value=None if pd.isna(vg) else float(vg))
        vergoeding.fill = _geel_vul
        vergoeding.number_format = EUR4
        verbruik = ws.cell(row=i, column=6,
                           value=None if pd.isna(vb) else float(vb))
        verbruik.fill = _geel_vul
        verbruik.number_format = KWH
        ws.cell(row=i, column=7, value="=(D{0}+E{0})*F{0}".format(i)).number_format = EUR

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:G{0}".format(len(d) + 1)
    return len(d)


def _blad_overzicht(wb: Workbook, maanden: list[str], n: int) -> None:
    ws = wb.create_sheet("Overzicht", 0)
    laatste = n + 1                       # laatste datarij in Kwartierdata
    K = "Kwartierdata!"
    mnd_bereik = K + "$B$2:$B$" + str(laatste)
    per_bereik = K + "$C$2:$C$" + str(laatste)
    prijs_bereik = K + "$D$2:$D$" + str(laatste)
    vbr_bereik = K + "$F$2:$F$" + str(laatste)
    kost_bereik = K + "$G$2:$G$" + str(laatste)

    ws["A1"] = "Eenheidsprijs per maand"
    ws["A1"].font = Font(bold=True, size=14, color=BLAUW)
    ws["A2"] = ("Plak je kwartierverbruiken in kolom F en je leveringsvergoeding in "
                "kolom E van het blad 'Kwartierdata'. De cijfers hieronder rekenen zichzelf.")
    ws["A2"].font = Font(italic=True, size=9, color="7F8C8D")
    ws["A3"] = ("Piek = weekdagen 08:00-19:45 · Dal = weekdagen 20:00-07:45 en weekends. "
                "Eenheidsprijs = totale kost / totaal verbruik, inclusief vergoeding.")
    ws["A3"].font = Font(italic=True, size=9, color="7F8C8D")

    kop = 5
    _kopregel(ws,
              ["Maand",
               "Verbruik totaal (kWh)", "Kost totaal (EUR)", "Eenheidsprijs (EUR/kWh)",
               "Verbruik piek (kWh)", "Kost piek (EUR)", "Eenheidsprijs piek (EUR/kWh)",
               "Verbruik dal (kWh)", "Kost dal (EUR)", "Eenheidsprijs dal (EUR/kWh)",
               "Baseload prijs (EUR/kWh)"],
              [11, 18, 16, 19, 17, 15, 18, 16, 15, 18, 17],
              rij=kop, hoogte=44)

    r = kop
    for r, maand in enumerate(maanden, start=kop + 1):
        m = "$A{0}".format(r)
        ws.cell(row=r, column=1, value=maand)
        ws.cell(row=r, column=2,
                value="=SUMIFS({0},{1},{2})".format(vbr_bereik, mnd_bereik, m)
                ).number_format = KWH
        ws.cell(row=r, column=3,
                value="=SUMIFS({0},{1},{2})".format(kost_bereik, mnd_bereik, m)
                ).number_format = EUR
        ws.cell(row=r, column=4,
                value='=IFERROR(C{0}/B{0},"")'.format(r)).number_format = EUR4
        ws.cell(row=r, column=5,
                value='=SUMIFS({0},{1},{2},{3},"{4}")'.format(
                    vbr_bereik, mnd_bereik, m, per_bereik, PIEK)).number_format = KWH
        ws.cell(row=r, column=6,
                value='=SUMIFS({0},{1},{2},{3},"{4}")'.format(
                    kost_bereik, mnd_bereik, m, per_bereik, PIEK)).number_format = EUR
        ws.cell(row=r, column=7,
                value='=IFERROR(F{0}/E{0},"")'.format(r)).number_format = EUR4
        ws.cell(row=r, column=8,
                value='=SUMIFS({0},{1},{2},{3},"{4}")'.format(
                    vbr_bereik, mnd_bereik, m, per_bereik, DAL)).number_format = KWH
        ws.cell(row=r, column=9,
                value='=SUMIFS({0},{1},{2},{3},"{4}")'.format(
                    kost_bereik, mnd_bereik, m, per_bereik, DAL)).number_format = EUR
        ws.cell(row=r, column=10,
                value='=IFERROR(I{0}/H{0},"")'.format(r)).number_format = EUR4
        ws.cell(row=r, column=11,
                value='=IFERROR(AVERAGEIF({0},{1},{2}),"")'.format(
                    mnd_bereik, m, prijs_bereik)).number_format = EUR4
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = _rand

    tot = r + 1
    eerste = kop + 1
    ws.cell(row=tot, column=1, value="Totaal")
    for c in (2, 3, 5, 6, 8, 9):
        L = get_column_letter(c)
        ws.cell(row=tot, column=c,
                value="=SUM({0}{1}:{0}{2})".format(L, eerste, r)
                ).number_format = KWH if c in (2, 5, 8) else EUR
    for doel, teller, noemer in ((4, "C", "B"), (7, "F", "E"), (10, "I", "H")):
        ws.cell(row=tot, column=doel,
                value='=IFERROR({0}{2}/{1}{2},"")'.format(teller, noemer, tot)
                ).number_format = EUR4
    ws.cell(row=tot, column=11,
            value='=IFERROR(AVERAGE({0}),"")'.format(prijs_bereik)).number_format = EUR4
    for c in range(1, 12):
        cel = ws.cell(row=tot, column=c)
        cel.font = Font(bold=True)
        cel.fill = _grijs_vul
        cel.border = _rand

    ws.freeze_panes = "A{0}".format(kop + 1)


def build_workbook(d: pd.DataFrame) -> bytes:
    """
    Bouwt de werkmap voor de eenheidsprijsberekening en geeft ze terug als
    bytes, klaar voor st.download_button.

    Verwacht een DataFrame met minstens de kolommen 'datetime' en 'price'.
    Bevat de DataFrame ook een kolom 'verbruik' en/of 'vergoeding', dan worden
    die gebruikt om kolom F resp. E van 'Kwartierdata' meteen voor te vullen
    i.p.v. leeg te laten.
    """
    if d is None or d.empty:
        raise ValueError("Geen data om in de werkmap te zetten.")
    for kolom in ("datetime", "price"):
        if kolom not in d.columns:
            raise ValueError("Kolom '{0}' ontbreekt in de data.".format(kolom))

    kolommen = ["datetime", "price"] + [k for k in ("verbruik", "vergoeding") if k in d.columns]
    d = d[kolommen].copy()
    d["datetime"] = pd.to_datetime(d["datetime"])
    d = d.dropna(subset=["datetime"]).sort_values("datetime").reset_index(drop=True)
    if d.empty:
        raise ValueError("Geen geldige tijdstempels in de data.")

    maanden = sorted(d["datetime"].dt.strftime("%Y-%m").unique().tolist())

    wb = Workbook()
    wb.remove(wb.active)
    _blad_overzicht(wb, maanden, _blad_kwartierdata(wb, d))
    wb.active = 0
    # openpyxl schrijft formules zonder gecachet resultaat; zonder deze vlag
    # tonen sommige viewers (en soms Excel zelf) de eenheidsprijs-kolommen
    # leeg totdat je handmatig herrekent (F9).
    wb.calculation.fullCalcOnLoad = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
